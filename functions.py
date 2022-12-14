from classes import *
import math
import openpyxl
import os
import pickle
from itertools import combinations
from scipy.spatial import distance
from tslearn.metrics import dtw, dtw_path, dtw_path_from_metric
from tslearn.barycenters import dtw_barycenter_averaging
import random
import matplotlib.pyplot as plt
import numpy

# return: vehdict, lanedict, col_info_list
def readtxt_VehData(fnum, varlist):
    f = open("data/" + str(fnum) + "_0vehicleData.txt", 'r')
    lines = f.readlines()
    lines.append("\n") #index 구분행 추가

    vehdict = {}
    lanedict = {}

    prev_index_veh_list = []
    current_index_veh_list = []
    collision_info_list = []

    indices = list(map(lambda x: (lines[13].split().index(x), x), varlist)) #varlist에 있는 각 variable의 위치

    for line in lines[14:]: #한 라인씩 읽으면서 각 vehicle object에 varlist의 항목들을 저장, lane에 시간별 차량 정보를 정렬하여 저장
        line = line.split()
        if len(line) > 1:
            #Vehicle object
            if line[2] not in vehdict:
                newveh = Vehicle(line[2])
                newveh.setVariableTypes(varlist)
                vehdict[line[2]] = newveh
            currentlineveh = vehdict[line[2]]
            for variable in indices:
                currentlineveh.appendValue(variable[1], line[variable[0]])

            current_index_veh_list.append(line[2])

            #Lane object
            if line[3] not in lanedict:
                newlane = Lane(line[3])
                lanedict[line[3]] = newlane
            lanedict[line[3]].appendVeh(float(line[1]), line[2], float(line[4]))


        else: #한 index가 끝났을 경우 충돌하여 사라진 차량 확인 후 collision 정보 저장
            for vehicle in prev_index_veh_list: #prev와 비교해서 사라진 차량 찾기
                if vehicle not in current_index_veh_list:
                    collision_vehicle = vehicle
                    col_veh_info = []
                    for variable in varlist:
                        col_veh_info.append(vehdict[collision_vehicle].getValues(variable)[-1])

                    time = vehdict[collision_vehicle].getValues("timeStamp")[-1]
                    difference = math.inf
                    closest_vehicle = None
                    closest_vehicle_timeindex = None
                    closest_veh_info = []
                    for vehid in prev_index_veh_list: #사라진 차량과 가장 근접한 차량 찾기
                        if vehid == collision_vehicle:
                            continue
                        timeindex = vehdict[vehid].getValues("timeStamp").index(time)
                        if vehdict[collision_vehicle].getValues("lane")[-1] == vehdict[vehid].getValues("lane")[timeindex]:
                            if abs(vehdict[collision_vehicle].getValues("lanepos")[-1] - vehdict[vehid].getValues("lanepos")[timeindex]) < difference:
                                difference = abs(vehdict[collision_vehicle].getValues("lanepos")[-1] - vehdict[vehid].getValues("lanepos")[timeindex])
                                closest_vehicle = vehid
                                closest_vehicle_timeindex = timeindex

                    for variable in varlist:
                        closest_veh_info.append(vehdict[closest_vehicle].getValues(variable)[closest_vehicle_timeindex])
                    collision_info_list.append((col_veh_info[0], col_veh_info[1], collision_vehicle, closest_vehicle))

            prev_index_veh_list = current_index_veh_list
            current_index_veh_list = []

    f.close()

    return vehdict, lanedict, collision_info_list

# return: platoon_history
def readtxt_InConfigData(fnum):
    f = open("data/" + str(fnum) + "_0plnConfig.txt", 'r')
    lines = f.readlines()
    lines.append("\n") #index 구분행 추가

    current_time = 0

    current_index_veh_list = []
    platoon_history = []
    for line in lines[14:]:
        if len(line) > 1:
            line = line.split()
            current_index_veh_list.append(line[2])
            current_time = float(line[lines[13].split().index("timestamp")])

        elif len(current_index_veh_list) > 0:
            if len(platoon_history) == 0:
                platoon_history.append((current_time, [current_index_veh_list]))
            elif platoon_history[-1][0] == current_time:
                platoon_history[-1][1].append(current_index_veh_list)
            else:
                platoon_history.append((current_time, [current_index_veh_list]))
            current_index_veh_list = []

    f.close()

    return platoon_history

# return: target_veh_platoon_history List((Time, List(plt_veh)))
def tracking_plt_history_for_target_veh(target_veh, structure_history):
    history = []
    start = 0
    end = 0
    cnt = -1
    last_ind = 0
    for tick in structure_history:
        cnt += 1
        for platoon in tick[1]:
            if target_veh in platoon:
                last_ind = cnt
                current = tick[0]
                if len(history) == 0:
                    history.append((current, platoon))
                    start = current
                    end = current
                elif history[-1][1] == platoon:
                    if end == start:
                        history.append((current, platoon))
                    else:
                        history[-1] = (current, platoon)
                    end = current
                else:
                    history.append((current, platoon))
                    start = current
                    end = current

    if len(history) != 0:
        last_platoon = history[-1][1][:]
        last_platoon.remove(target_veh)

        #Leave 확인
        if last_ind != len(structure_history) - 1:
            for tick in structure_history[last_ind+1:]:
                for platoon in tick[1]:
                    for veh in last_platoon:
                        if veh in platoon:
                            history.append((tick[0], "Leave"))
                            return history

    return history

# Calculate the order distance in the platoon vehicles
def structure_distance(base_veh, target_veh, time, structure_history):
    prev_history = []
    for change_point in structure_history:
        if time >= change_point[0]:
            prev_history = change_point[1]
        else:
            break
    if prev_history == "Leave":
        return -1
    elif target_veh in prev_history:
        return abs(prev_history.index(base_veh) - prev_history.index(target_veh))
    else:
        return -1

# Calculate the order distance in the lane
def lane_distance(base_veh, target_veh, time, lane_history):
    prev_history = None
    for tick in lane_history:
        if time >= tick[0]:
            prev_history = tick[1]
        else:
            break

    if target_veh in prev_history:
        return abs(prev_history.index(base_veh) - prev_history.index(target_veh))
    else:
        return -1

# Calculate the position distance in the lane
def position_distance(base_veh, target_veh, time, lane_history):
    prev_history = None
    for tick in lane_history:
        if time >= tick[0]:
            prev_history = tick[1]
        else:
            break

    veh_list = list(map(lambda x: x[0], prev_history))
    if target_veh in veh_list:
        return abs(prev_history[veh_list.index(base_veh)][1] - prev_history[veh_list.index(target_veh)][1])
    else:
        return -1

# Find front vehs based on lane_history at the specific time
def get_front_veh(base_veh, time, lane):
    lane_info = lane.getperiodinfo(time, time)[0]
    print("Lane information: ", lane_info)

    front_veh_list = []
    for i in range(len(lane_info[1])):
        if lane_info[1][i][0] == base_veh:
            front_veh_list = lane_info[1][i+1:]
            break

    print("Front vehicle list: ", front_veh_list)
    return front_veh_list

# Calculate the structure, lane composite distance
def structure_lane_composite_distance(base_veh, time, ordered_veh_list, structure_history):
    ordered_veh_numbering = []
    cnt = 0
    for veh in ordered_veh_list:
        current_veh_structure_history = tracking_plt_history_for_target_veh(base_veh, structure_history)
        str_dist = structure_distance(base_veh, veh[0], time, current_veh_structure_history)
        if str_dist == -1:
            cnt += 1
        ordered_veh_numbering.append(cnt)
        base_veh = veh[0]

    print("Front vehicle numbering: ", ordered_veh_numbering)

    return ordered_veh_numbering

# Select the vehicles that might be related to the collision
def select_related_vehicles(front_veh_list, front_veh_numbering, compositionbound, distancebound):
    veh_list = front_veh_list
    veh_numbering = front_veh_numbering

    for i in range(len(front_veh_numbering) - 1):
        if front_veh_numbering[i+1] != front_veh_numbering[i] and front_veh_list[i+1][1] - front_veh_list[i][1] > distancebound:
            veh_list = front_veh_list[:i+1]
            veh_numbering = front_veh_numbering[:i+1]
            break

    try:
        ind = veh_numbering.index(compositionbound + 1)
        return veh_list[:ind], veh_numbering[:ind]
    except ValueError:
        return veh_list, veh_numbering

# Generate group objects and return list of group objects
def generate_group_object(col_veh, veh_list, veh_numbering, vehdict, lanedict, timescope, col_lane):
    groups = []
    col_veh = Group([col_veh])
    groups.append(col_veh)

    current_numbering = veh_numbering[0]
    current_list = []
    for i in range(len(veh_list)):
        if current_numbering == veh_numbering[i]:
            current_list.append(vehdict[veh_list[i][0]])
        else:
            new_group = Group(current_list)
            groups.append(new_group)
            current_numbering = veh_numbering[i]
            current_list = [vehdict[veh_list[i][0]]]
    new_group = Group(current_list)
    groups.append(new_group)


    for i in range(len(groups)):
        group = groups[i]
        if len(group.vehlist) == 1:
            veh = group.vehlist[0]
            start_ind = veh.vardict["timeStamp"].index(timescope[0])
            end_ind = veh.vardict["timeStamp"].index(timescope[1])
            group.timestamp = veh.getValues("timeStamp")[start_ind:end_ind + 1]
            group.lane = veh.getValues("lane")[start_ind:end_ind + 1]
            group.avg_occupancy = [0 for i in range(end_ind - start_ind + 1)]
            group.speed = veh.getValues("speed")[start_ind:end_ind + 1]

            for i in range(len(group.timestamp)):
                lane_info = lanedict[group.lane[i]].getperiodinfo(group.timestamp[i], group.timestamp[i])[0][1]
                lane_vehs = list(map(lambda x: x[0], lane_info))

                veh_index = lane_vehs.index(veh.vehid)
                if veh_index != len(lane_vehs) - 1:
                    group.distance.append(round(lane_info[veh_index + 1][1] - lane_info[veh_index][1], 2))
                else:
                    group.distance.append(math.inf)

            group.lane = list(map(lambda x: 0 if x == col_lane.lanenum else 1, veh.getValues("lane")[start_ind:end_ind + 1]))


        else:
            back_veh = group.vehlist[0]
            front_veh = group.vehlist[-1]
            start_ind = front_veh.vardict["timeStamp"].index(timescope[0])
            end_ind = front_veh.vardict["timeStamp"].index(timescope[1])
            group.timestamp = front_veh.getValues("timeStamp")[start_ind:end_ind + 1]
            group.lane = front_veh.getValues("lane")[start_ind:end_ind + 1]
            group.speed = back_veh.getValues("speed")[start_ind:end_ind + 1]

            for i in range(len(group.timestamp)):
                lane_info = lanedict[group.lane[i]].getperiodinfo(group.timestamp[i], group.timestamp[i])[0][1]
                lane_vehs = list(map(lambda x: x[0], lane_info))

                front_veh_index = lane_vehs.index(front_veh.vehid)
                back_veh_index = lane_vehs.index(back_veh.vehid)
                group.avg_occupancy.append(round((lane_info[front_veh_index][1] - lane_info[back_veh_index][1])/len(group.vehlist), 2))
                if front_veh_index != len(lane_vehs) - 1:
                    group.distance.append(round(lane_info[front_veh_index + 1][1] - lane_info[front_veh_index][1], 2))
                else:
                    group.distance.append(math.inf)

            group.lane = list(map(lambda x: 0 if x == col_lane.lanenum else 1, front_veh.getValues("lane")[start_ind:end_ind + 1]))

    return groups

# Generating excel files
def write_group_object_to_xlsx(fnum, groups):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "Vehicle List"
    for i in range(len(groups)):
        sheet.cell(row=2, column = i+1).value = ', '.join(veh.vehid for veh in groups[i].vehlist)

    sheet.append([])
    indexlist = ['timestamp']
    for i in range(len(groups)):
        indexlist += ['group' + str(i) + '_speed', 'group' + str(i) + '_distance', 'group' + str(i) + '_lane', 'group' + str(i) + '_avg_occupancy']
    sheet.append(indexlist)

    for i in range(len(groups[0].timestamp)):
        current_time_info = [groups[0].timestamp[i]]
        for group in groups:
            current_time_info += [group.speed[i], 'inf' if group.distance[i] == math.inf else group.distance[i], group.lane[i], group.avg_occupancy[i]]
        sheet.append(current_time_info)

    wb.save('groupdata/' + str(fnum) + '_groupData.xlsx')

# Generating excel files that contains group data from initail raw data
def make_group_xlsx_from_raw_data(filename, compositionbound, distancebound, timescope):
    wb = openpyxl.load_workbook("data/" + filename, data_only = True)
    sheet = wb['Falselog']
    col_file_num = []
    for row in list(sheet.rows)[1:]:
        filenum = row[0].value[0:4]
        print(filenum)
        col_file_num.append(filenum)
        row_num = row[0].row

        try:
            print("filenum: ", filenum)
            vehdict, lanedict, col_info = readtxt_VehData(filenum, ["timeStamp", "lane", "lanepos", "speed", "accel"])
            print("Collision Information: ", col_info)
            if len(col_info) != 0:
                structure_history = readtxt_InConfigData(filenum)

                # 두 충돌 차량의 위치를 바탕으로 뒤에 있는 차량 찾기
                first_col = col_info[0]
                col_time = first_col[0]
                col_lane = lanedict[first_col[1]]
                col_veh1 = vehdict[first_col[2]]
                col_veh2 = vehdict[first_col[3]]
                sheet.cell(row=row_num, column=7).value = first_col[2]
                sheet.cell(row=row_num, column=8).value = first_col[3]
                col_veh1_pos = col_veh1.getValues("lanepos", col_time)
                col_veh2_pos = col_veh2.getValues("lanepos", col_time)

                behind_veh = col_veh1
                if col_veh1_pos > col_veh2_pos:
                    behind_veh = col_veh2

                front_veh_list = get_front_veh(behind_veh.vehid, col_time, col_lane)

                front_veh_numbering = structure_lane_composite_distance(behind_veh.vehid, col_time, front_veh_list, structure_history)

                col_related_veh_list, col_related_veh_numbering = select_related_vehicles(front_veh_list, front_veh_numbering, compositionbound, distancebound)
                print(col_related_veh_list)

                plt_relation = plt_relation_check(behind_veh.vehid, col_related_veh_list)
                sheet.cell(row=row_num, column=9).value = plt_relation

                if plt_relation:
                    groups = generate_group_object(behind_veh, col_related_veh_list, col_related_veh_numbering, vehdict, lanedict, (round(col_time - timescope, 1), col_time), col_lane)

                    write_group_object_to_xlsx(filenum, groups)

        except FileNotFoundError:
            continue
    wb.save("data/" + filename)

# Read group excel files and generate MTS
def read_group_xlsx():
    path = 'groupdata/'
    dir_list = [file for file in os.listdir(path) if file.endswith('.xlsx')]

    filelist = []
    mts_dataset = []
    for file in dir_list:
        print(file)
        filelist.append(file[0:4])
        sheet = openpyxl.load_workbook(path + file, data_only = True)['Sheet']
        rows = list(sheet.rows)

        group_list = [group for group in list(map(lambda x: x.value, rows[1])) if group != None]
        group_num = len(group_list)

        mts = []
        for line in rows[4:]:
            mts.append([value for value in list(map(lambda x: x.value, line[1:])) if value != None])

        mts_dataset.append(mts)

    with open('groupdata/mtsdata.pickle', 'wb') as f:
        pickle.dump((filelist, mts_dataset), f)

    return filelist, mts_dataset

def preprocess(mts_dataset, distancebound):
    speed_values = []
    distance_values = []
    occupancy_ratio_values = []
    for mts in mts_dataset:
        group_num = len(mts[0])//4
        for i in range(group_num):
            speed_values += list(map(lambda x: x[4*i], mts))
            distance_values += [value for value in list(map(lambda x: x[4*i+1], mts)) if value != "inf"]
            occupancy_ratio_values += list(map(lambda x: x[4*i+3], mts))

    min_speed = min(speed_values)
    max_speed = max(speed_values)
    min_distance = min(distance_values)
    max_distance = max(distance_values)
    max_distance = distancebound if max_distance > distancebound else max_distance
    min_avg_occupancy = min(occupancy_ratio_values)
    max_avg_occupancy = max(occupancy_ratio_values)
    avg_occupancy_bound = 15
    max_avg_occupancy = avg_occupancy_bound if max_avg_occupancy > avg_occupancy_bound else max_avg_occupancy

    for mts in mts_dataset:
        group_num = len(mts[0])//4
        for tick in mts:
            for i in range(group_num):
                tick[4*i] = (tick[4*i] - min_speed)/(max_speed - min_speed)
                if tick[4*i+3] > avg_occupancy_bound:
                    tick[4*i+3] = 1
                else:
                    tick[4*i+3] = (tick[4*i+3] - min_avg_occupancy) / (max_avg_occupancy - min_avg_occupancy)
                if tick[4*i+1] == "inf" or tick[4*i+1] > distancebound:
                    tick[4*i+1] = 1
                else:
                    tick[4*i+1] = (tick[4*i+1] - min_distance)/(max_distance - min_distance)

    return mts_dataset

def preprocess_znorm(mts_dataset, distancebound):
    speed_values = []
    distance_values = []
    lane_values = []
    occupancy_ratio_values = []

    for mts in mts_dataset:
        group_num = len(mts[0])//4
        for i in range(group_num):
            speed_values += list(map(lambda x: x[4*i], mts))
            distance_values += [min(value, distancebound) for value in list(map(lambda x: x[4*i+1], mts))]
            lane_values += list(map(lambda x: x[4*i+2], mts))
            occupancy_ratio_values += list(map(lambda x: x[4*i+3], mts))

    mean_speed = numpy.mean(speed_values)
    mean_distance = numpy.mean(distance_values)
    mean_lane = numpy.mean(lane_values)
    mean_occupancy_ratio = numpy.mean(occupancy_ratio_values)
    std_speed = numpy.std(speed_values)
    std_distance = numpy.std(distance_values)
    std_lane = numpy.std(lane_values)
    std_occupancy_ratio = numpy.std(occupancy_ratio_values)

    for mts in mts_dataset:
        group_num = len(mts[0])//4
        for tick in mts:
            for i in range(group_num):
                tick[4*i] = (tick[4*i] - mean_speed)/std_speed
                tick[4*i+1] = (tick[4*i+1] - mean_distance)/std_distance
                tick[4*i+2] = (tick[4*i+2] - mean_lane)/std_lane
                tick[4*i+3] = (tick[4*i+3] - mean_occupancy_ratio)/std_occupancy_ratio

    return mts_dataset


# Calculating two MTS with different dimensions
def calculate_distance_btw_two_MTS(mts1, mts2):
    dim1 = len(mts1[0])//4
    dim2 = len(mts2[0])//4
    if dim1 < dim2:
        mts1, mts2 = mts2, mts1
        dim1, dim2 = dim2, dim1

    matching_combination = [[0] + list(x) for x in combinations(range(1,dim1),dim2-1)]
    min_distance = math.inf
    min_matching = None
    for matching_case in matching_combination:
        cutted_mts = []
        for tick in mts1:
            cutted_tick = []
            for group_num in matching_case:
                cutted_tick += tick[4*group_num: 4*(group_num+1)]
            cutted_mts.append(cutted_tick)
        distance = mydtw(cutted_mts, mts2)
        if distance < min_distance:
            min_distance = distance
            min_matching = matching_case

    return min_distance, min_matching

# DTW calculation btw 2 MTS with same parameter numbers
def mydtw(mts1, mts2):
    array = [[math.inf for j in range(len(mts2)+1)] for i in range(len(mts1)+1)]
    array[0][0] = 0

    for i in range(1,len(mts1)+1):
        for j in range(1,len(mts2)+1):
            cost = distance.euclidean(mts1[i-1], mts2[j-1])
            array[i][j] = cost + min(array[i-1][j],array[i][j-1],array[i-1][j-1])

    return array[len(mts1)][len(mts2)]

# Return center of mtslist
def get_barycenter(mtslist, compositionbound):
    max_dim = 0
    max_dim_position = None
    for i in range(len(mtslist)):
        if len(mtslist[i][0])//4 > max_dim:
            max_dim = len(mtslist[i][0])//4
            max_dim_position = [i]
        elif len(mtslist[i][0])//4 == max_dim:
            max_dim_position.append(i)

    random_max_dim_position = random.choice(max_dim_position)

    matching_list = []
    for i in range(len(mtslist)):
        distance, matching = calculate_distance_btw_two_MTS(mtslist[random_max_dim_position], mtslist[i])
        matching_list.append(matching)

    barycenter = [[] for i in range(len(mtslist[0]))]
    for i in range(compositionbound+1):
        dataset = []
        for j in range(len(mtslist)):
            if i in matching_list[j]:
                ind = matching_list[j].index(i)
                dataset.append(list(map(lambda x: x[4*ind:4*(ind+1)], mtslist[j])))
        if len(dataset) != 0:
            current_composition_distance_center = dtw_barycenter_averaging(dataset, barycenter_size=len(mtslist[0])).tolist()
            for j in range(len(mtslist[0])):
                barycenter[j] += current_composition_distance_center[j]

    return barycenter

# Retrun clustered result
def dtwKMeans(mtslist, numcenter, compositionbound):
    #initial centers
    centers = random.sample(mtslist, numcenter)

    errors = [math.inf for i in range(numcenter)]
    nearest_center_index = [None for i in range(len(mtslist))]

    while len([error for error in errors if error != 0]) >0:
        for i in range(len(mtslist)):
            min_index = None
            min_distance = math.inf
            for j in range(len(centers)):
                distance = calculate_distance_btw_two_MTS(centers[j], mtslist[i])[0]
                if distance < min_distance:
                    min_index = j
                    min_distance = distance
            nearest_center_index[i] = min_index
        print(nearest_center_index)
        #Calculate barycenters
        for i in range(numcenter):
            cluster_points = [mtslist[j] for j in range(len(mtslist)) if nearest_center_index[j] == i]
            if len(cluster_points) != 0:
                new_center = get_barycenter(cluster_points, compositionbound)
                errors[i] = calculate_distance_btw_two_MTS(centers[i], new_center)[0]
                centers[i] = new_center
            else:
                errors[i] = 0

        print("Current error: ", errors)
        print("Current cluster:", nearest_center_index)
    with open('outputdata/clusterdata.pickle', 'wb') as f:
        pickle.dump((nearest_center_index, centers), f)

    return nearest_center_index, centers

#
def visualization(mtslist, cluster_result, centers):
    for i in range(len(centers)):
        tick = list(range(len(centers[i])))

        for j in range(len(cluster_result)):
            if cluster_result[j] == i:
                matching = calculate_distance_btw_two_MTS(centers[i], mtslist[j])[1]
                for k in matching:
                    for m in range(4):
                        plt.subplot(len(centers[i][0]), 1, 4*k+m+1)
                        ind = matching.index(k)
                        value = [x[4*ind+m] for x in mtslist[j]]
                        plt.ylim(0,1)
                        plt.plot(tick, value, linewidth = 0.5)

        for j in range(len(centers[i][0])):
            plt.subplot(len(centers[i][0]), 1, j+1)
            value = [x[j] for x in centers[i]]
            plt.ylim(0,1)
            plt.plot(tick, value, 'r', linewidth= 2)

        plt.savefig('outputdata/cluster_' + str(i) + '_mts_figure.png', dpi = 400)
        plt.clf()
        #plt.show()


# return: Collision type
def plt_relation_check(behind_vehid, front_veh_list):
    if 'veh' in behind_vehid:
        return True
    else:
        front_vehid_list = list(map(lambda x: x[0], front_veh_list))
        print(front_vehid_list)
        for veh in front_vehid_list:
            if 'veh' in veh:
                return True
        return False

# return: time period that should be analyzed
def getTimeScope(col_time, plt_history):
    for i in range(len(plt_history) - 1):
        if plt_history[i][0] < col_time <= plt_history[i + 1][0]:
            return [plt_history[i][0], col_time]





